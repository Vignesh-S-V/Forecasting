import warnings
warnings.filterwarnings("ignore")

import io
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from scipy import stats
from scipy.optimize import minimize, differential_evolution
from scipy.signal import savgol_filter
import streamlit as st

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Textile Sales Forecasting Pro",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;600&display=swap');
html, body, [class*="css"] { font-family: 'Space Grotesk', sans-serif; }
.stApp { background: linear-gradient(135deg, #0f0c29, #1a1a4e, #0f0c29); color: #e8e8f0; }
[data-testid="stSidebar"] { background: linear-gradient(180deg, #13103a 0%, #1e1a5e 100%); border-right: 1px solid rgba(255,255,255,0.08); }
[data-testid="stSidebar"] * { color: #d4d4f0 !important; }
.metric-card { background: rgba(255,255,255,0.05); border: 1px solid rgba(255,255,255,0.12); border-radius: 16px; padding: 20px 24px; backdrop-filter: blur(10px); transition: all 0.3s ease; }
.metric-card:hover { border-color: rgba(120,100,255,0.4); background: rgba(255,255,255,0.08); }
.metric-label { font-size: 12px; font-weight: 600; letter-spacing: 1.5px; text-transform: uppercase; color: #8888cc; margin-bottom: 8px; }
.metric-value { font-size: 28px; font-weight: 700; color: #ffffff; font-family: 'JetBrains Mono', monospace; }
.metric-sub { font-size: 12px; color: #6666aa; margin-top: 4px; }
.section-header { font-size: 20px; font-weight: 700; color: #a09cf7; border-left: 4px solid #7c6fee; padding-left: 14px; margin: 28px 0 18px 0; letter-spacing: 0.5px; }
[data-testid="stFileUploader"] { background: rgba(120,100,255,0.08) !important; border: 2px dashed rgba(120,100,255,0.4) !important; border-radius: 14px !important; padding: 20px !important; }
.stButton > button { background: linear-gradient(135deg, #7c6fee, #9b8cf5) !important; color: white !important; border: none !important; border-radius: 10px !important; font-weight: 600 !important; font-size: 15px !important; padding: 10px 28px !important; transition: all 0.3s ease !important; }
.stButton > button:hover { transform: translateY(-2px); box-shadow: 0 8px 24px rgba(124,111,238,0.5) !important; }
.stSelectbox [data-baseweb="select"] { background: rgba(255,255,255,0.06) !important; border-color: rgba(255,255,255,0.15) !important; border-radius: 10px !important; }
.stTabs [data-baseweb="tab-list"] { background: rgba(255,255,255,0.04) !important; border-radius: 12px !important; gap: 4px; padding: 4px; }
.stTabs [data-baseweb="tab"] { border-radius: 9px !important; color: #9999cc !important; font-weight: 600 !important; }
.stTabs [aria-selected="true"] { background: rgba(124,111,238,0.35) !important; color: #ffffff !important; }
.banner-warn { background: linear-gradient(135deg, rgba(255,140,0,0.15), rgba(255,80,80,0.10)); border: 1px solid rgba(255,140,0,0.45); border-radius: 14px; padding: 16px 22px; margin: 12px 0 18px 0; }
.banner-good { background: linear-gradient(135deg, rgba(80,200,120,0.12), rgba(80,180,255,0.08)); border: 1px solid rgba(80,200,120,0.40); border-radius: 14px; padding: 16px 22px; margin: 12px 0 18px 0; }
.banner-title { font-size: 15px; font-weight: 700; margin-bottom: 6px; }
.banner-body { font-size: 13px; line-height: 1.6; }
.warn-title { color: #ffaa44; } .warn-body { color: #ccaa77; }
.good-title { color: #50d090; } .good-body { color: #88ccaa; }
.acc-high { display:inline-block; background:rgba(80,200,120,0.18); border:1px solid rgba(80,200,120,0.45); color:#50d090; border-radius:8px; padding:4px 14px; font-size:13px; font-weight:700; }
.acc-mid  { display:inline-block; background:rgba(255,165,0,0.18);  border:1px solid rgba(255,165,0,0.45);  color:#ffaa44; border-radius:8px; padding:4px 14px; font-size:13px; font-weight:700; }
.acc-low  { display:inline-block; background:rgba(255,80,80,0.18);  border:1px solid rgba(255,80,80,0.45);  color:#ff8080; border-radius:8px; padding:4px 14px; font-size:13px; font-weight:700; }
</style>
""", unsafe_allow_html=True)

PLOTLY_THEME = dict(
    template="plotly_dark",
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(255,255,255,0.03)",
    font=dict(family="Space Grotesk", color="#d4d4f0"),
    xaxis=dict(gridcolor="rgba(255,255,255,0.06)", showgrid=True),
    yaxis=dict(gridcolor="rgba(255,255,255,0.06)", showgrid=True),
    margin=dict(l=60, r=40, t=60, b=50),
)
ACCENT = ["#7c6fee", "#f5a623", "#50d090", "#e05cf5", "#f55050", "#50c8f5"]

MIN_MONTHS_FULL = 18
MIN_MONTHS_BASIC = 6

# ─────────────────────────────────────────────────────────────────────────────
# DATA UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_excel(file_bytes):
    return pd.read_excel(io.BytesIO(file_bytes))

def find_columns(df):
    col_map = {}; mapped = set()
    for c in df.columns:
        cl = str(c).strip().lower()
        if cl == "date" and "Date" not in mapped:
            col_map[c] = "Date"; mapped.add("Date")
        elif cl in ("cy value","cyvalue","cy_value","sales","value","amount","revenue") and "Sales" not in mapped:
            col_map[c] = "Sales"; mapped.add("Sales")
        elif cl in ("subgrp","sub_grp","sub grp","productname","product_name","product name","product","category","subgroup") and "SUBGRP" not in mapped:
            col_map[c] = "SUBGRP"; mapped.add("SUBGRP")
    return col_map

def safe_series(df, col):
    r = df[col]
    return r.iloc[:, 0] if isinstance(r, pd.DataFrame) else r

def dedup_cols(df):
    seen = {}; cols = []
    for c in df.columns:
        if c not in seen: seen[c] = 0; cols.append(c)
        else: seen[c] += 1; cols.append(f"{c}__d{seen[c]}")
    df.columns = cols; return df

def data_status(n):
    if n < MIN_MONTHS_BASIC:
        return dict(ok=False, tier="bad", sl=max(2,n//2),
                    msg=f"❌ Only {n} months — need ≥{MIN_MONTHS_BASIC}.")
    elif n < MIN_MONTHS_FULL:
        return dict(ok=True, tier="limited", sl=min(n//2,6),
                    msg=f"⚠️ {n} months — limited accuracy. Upload ≥18 months for full seasonal capture.")
    else:
        return dict(ok=True, tier="full", sl=12,
                    msg=f"✅ {n} months — full accuracy mode active.")

def acc_badge(mape):
    if mape < 5:   return '<span class="acc-high">🎯 Excellent — MAPE &lt;5%</span>'
    elif mape < 10: return '<span class="acc-high">✅ Good — MAPE &lt;10%</span>'
    elif mape < 15: return '<span class="acc-mid">⚠️ Fair — MAPE 10–15%</span>'
    else:           return '<span class="acc-low">❌ Poor — MAPE &gt;15%</span>'

# ─────────────────────────────────────────────────────────────────────────────
# OUTLIER DETECTION  (conservative — don't kill real signals)
# ─────────────────────────────────────────────────────────────────────────────

def detect_outliers(series: pd.Series):
    """
    Uses Hampel identifier (sliding window MAD) instead of global IQR.
    Much more conservative — only flags extreme spikes, not genuine peaks.
    """
    y = series.values.astype(float); n = len(y)
    window = max(5, min(7, n // 3))
    flags = np.zeros(n, dtype=bool)
    for i in range(n):
        lo = max(0, i - window // 2); hi = min(n, i + window // 2 + 1)
        neighbors = np.concatenate([y[lo:i], y[i+1:hi]])
        if len(neighbors) < 3: continue
        med = np.median(neighbors); mad = np.median(np.abs(neighbors - med))
        sigma_hat = 1.4826 * mad
        if sigma_hat > 0 and abs(y[i] - med) > 4.0 * sigma_hat:
            flags[i] = True
    y_clean = y.copy()
    for i in np.where(flags)[0]:
        left  = next((j for j in range(i-1,-1,-1) if not flags[j]), None)
        right = next((j for j in range(i+1,n)     if not flags[j]), None)
        if left is not None and right is not None:
            t = (i - left) / (right - left); y_clean[i] = y[left]*(1-t) + y[right]*t
        elif left  is not None: y_clean[i] = y[left]
        elif right is not None: y_clean[i] = y[right]
    return pd.Series(y_clean, index=series.index), flags, {"n_outliers": int(flags.sum()), "idx": list(np.where(flags)[0])}

# ─────────────────────────────────────────────────────────────────────────────
# STL DECOMPOSITION  (proper iterative LOESS-based)
# ─────────────────────────────────────────────────────────────────────────────

def stl_decompose(y: np.ndarray, period: int = 12, robust: bool = True):
    """
    Simplified but correct STL: iterative trend + seasonal smoothing.
    Returns trend, seasonal, residual components.
    """
    n = len(y); period = min(period, n // 2)
    if period < 2: period = 2

    # --- Initial trend via Savitzky-Golay or moving average ---
    wlen = min(2 * period + 1, n if n % 2 == 1 else n - 1)
    wlen = max(wlen, 3); wlen = wlen if wlen % 2 == 1 else wlen - 1
    poly = min(2, wlen - 1)
    try:
        trend = savgol_filter(y, wlen, poly)
    except Exception:
        trend = np.convolve(y, np.ones(period)/period, mode='same')

    for _ in range(5):  # STL outer iterations
        detrended = y - trend

        # Compute seasonal component by averaging same-phase points
        seasonal = np.zeros(n)
        for p in range(period):
            idx = np.arange(p, n, period)
            vals = detrended[idx]
            if robust:
                # Bisquare robustness weights
                med = np.median(vals); mad = max(np.median(np.abs(vals - med)), 1e-8)
                u = np.abs(vals - med) / (6 * mad)
                w = np.where(u < 1, (1 - u**2)**2, 0.0)
                w_sum = w.sum()
                s_val = (w * vals).sum() / w_sum if w_sum > 0 else np.mean(vals)
            else:
                s_val = np.mean(vals)
            seasonal[idx] = s_val

        # Remove mean from seasonal so it sums to 0
        seasonal -= seasonal.mean()

        # Re-estimate trend from seasonally adjusted series
        sa = y - seasonal
        try:
            trend = savgol_filter(sa, wlen, poly)
        except Exception:
            trend = np.convolve(sa, np.ones(period)/period, mode='same')

    residual = y - trend - seasonal
    return trend, seasonal, residual

# ─────────────────────────────────────────────────────────────────────────────
# MODEL 1: STL + DAMPED TREND (best for trending seasonal data)
# ─────────────────────────────────────────────────────────────────────────────

def model_stl_damped(y: np.ndarray, period: int, fp: int):
    """
    STL decomposition → forecast trend with damped Holt + replay seasonal pattern.
    """
    n = len(y)
    trend, seasonal, residual = stl_decompose(y, period=period, robust=True)

    # Fit damped Holt's method on trend
    def sse(params):
        alpha, beta, phi = params
        if not (0 < alpha < 1 and 0 < beta < 0.5 and 0.8 <= phi <= 1.0): return 1e15
        L, T = trend[0], trend[1] - trend[0]
        err_sq = 0.0
        for t in range(1, len(trend)):
            forecast_t = L + phi * T
            e = trend[t] - forecast_t
            err_sq += e ** 2
            L = alpha * trend[t] + (1 - alpha) * (L + phi * T)
            T = beta * (L - (L - alpha*(trend[t] - (L + phi*T)) )) + (1 - beta) * phi * T
        return err_sq

    try:
        res = minimize(sse, [0.4, 0.1, 0.95], method="Nelder-Mead",
                       options={"maxiter": 5000, "xatol": 1e-7, "fatol": 1e-7})
        alpha, beta, phi = np.clip(res.x, [0.01, 0.01, 0.8], [0.99, 0.49, 1.0])
    except Exception:
        alpha, beta, phi = 0.4, 0.1, 0.95

    # Re-run Holt on trend with best params to get final L, T
    L, T = trend[0], trend[1] - trend[0]
    for t in range(1, len(trend)):
        L_new = alpha * trend[t] + (1 - alpha) * (L + phi * T)
        T_new = beta * (L_new - L) + (1 - beta) * phi * T
        L, T = L_new, T_new

    # Trend forecast
    trend_fc = np.array([L + sum(phi**h for h in range(1, h+1)) * T for h in range(1, fp+1)])

    # Seasonal forecast: repeat last full cycle
    seasonal_fc = np.array([seasonal[n - period + (h-1) % period] for h in range(1, fp+1)])

    # Fitted
    trend_fit = trend.copy()
    seasonal_fit = seasonal.copy()
    fitted = np.maximum(trend_fit + seasonal_fit, 0)
    forecast = np.maximum(trend_fc + seasonal_fc, 0)

    mape = float(np.mean(np.abs((y - fitted) / (np.abs(y) + 1e-9))) * 100)
    return {"name": "STL + Damped Holt", "fitted": fitted.tolist(),
            "forecast": forecast.tolist(), "mape": mape,
            "components": {"trend": trend, "seasonal": seasonal, "residual": residual}}

# ─────────────────────────────────────────────────────────────────────────────
# MODEL 2: OPTIMISED HOLT-WINTERS ETS (additive, multiplicative tried)
# ─────────────────────────────────────────────────────────────────────────────

def model_holtwinters(y: np.ndarray, period: int, fp: int):
    n = len(y); period = min(period, n // 2) if n >= period * 2 else max(2, n // 3)

    def run_hw(y_, alpha, beta, gamma, mul_season=False):
        n_ = len(y_)
        if n_ < 2 * period:
            L = np.mean(y_[:max(1, n_//2)]); T = 0.0
            S = list(y_[:period] - L) if not mul_season else list(y_[:period] / (L + 1e-9))
        else:
            L = np.mean(y_[:period])
            T = (np.mean(y_[period:2*period]) - np.mean(y_[:period])) / period
            if mul_season:
                S = list(y_[:period] / (L + 1e-9))
            else:
                S = list(y_[:period] - L)
        fitted_ = []
        for t in range(n_):
            s = S[t % period]
            f = (L + T) * s if mul_season else L + T + s
            fitted_.append(max(f, 0))
            if mul_season:
                L_new = alpha * (y_[t] / (s + 1e-9)) + (1 - alpha) * (L + T)
            else:
                L_new = alpha * (y_[t] - s) + (1 - alpha) * (L + T)
            T_new = beta * (L_new - L) + (1 - beta) * T
            if mul_season:
                S[t % period] = gamma * (y_[t] / (L_new + 1e-9)) + (1 - gamma) * s
            else:
                S[t % period] = gamma * (y_[t] - L_new) + (1 - gamma) * s
            L, T = L_new, T_new
        fc_ = []
        for h in range(1, fp + 1):
            s = S[(n_ - 1 + h) % period]
            fc_.append(max((L + h * T) * s if mul_season else L + h * T + s, 0))
        return fitted_, fc_, L, T, S

    best_res = None; best_sse = np.inf
    for mul in [False, True]:
        def sse_fn(p):
            a, b, g = p
            if not (0 < a < 1 and 0 < b < 0.3 and 0 < g < 1): return 1e15
            ft, _, _, _, _ = run_hw(y, a, b, g, mul)
            return float(np.sum((y - np.array(ft))**2))
        try:
            r = minimize(sse_fn, [0.3, 0.05, 0.3], method="Nelder-Mead",
                         options={"maxiter": 10000, "xatol": 1e-8})
            if r.fun < best_sse:
                best_sse = r.fun
                best_res = (np.clip(r.x, 0.01, 0.99), mul)
        except Exception:
            pass

    if best_res is None:
        best_res = ([0.3, 0.05, 0.3], False)

    (alpha, beta, gamma), mul_s = best_res
    fitted, forecast, _, _, _ = run_hw(y, alpha, beta, gamma, mul_s)
    mape = float(np.mean(np.abs((y - np.array(fitted)) / (np.abs(y) + 1e-9))) * 100)
    return {"name": "Holt-Winters ETS (opt.)", "fitted": fitted,
            "forecast": forecast, "mape": mape, "params": {}}

# ─────────────────────────────────────────────────────────────────────────────
# MODEL 3: SEASONAL NAIVE + DRIFT (strong baseline for seasonal data)
# ─────────────────────────────────────────────────────────────────────────────

def model_snaive_drift(y: np.ndarray, period: int, fp: int):
    n = len(y); period = min(period, n - 1)
    # Compute average drift over last 2 full seasonal cycles
    n_cycles = min(2, n // period)
    if n_cycles >= 1:
        start = n - n_cycles * period
        drift_per_period = (y[-1] - y[start]) / (n_cycles * period) if n_cycles * period > 0 else 0
    else:
        drift_per_period = (y[-1] - y[0]) / max(n - 1, 1)

    fitted = []
    for t in range(n):
        prev_season = y[t - period] if t >= period else y[t % period]
        fitted.append(max(prev_season + drift_per_period, 0))

    forecast = []
    for h in range(1, fp + 1):
        prev_s_idx = n - period + (h - 1) % period
        if 0 <= prev_s_idx < n:
            forecast.append(max(y[prev_s_idx] + drift_per_period * h, 0))
        else:
            forecast.append(max(y[-1] + drift_per_period * h, 0))

    fitted_arr = np.array(fitted)
    mape = float(np.mean(np.abs((y - fitted_arr) / (np.abs(y) + 1e-9))) * 100)
    return {"name": "Seasonal Naïve + Drift", "fitted": fitted,
            "forecast": forecast, "mape": mape, "params": {}}

# ─────────────────────────────────────────────────────────────────────────────
# MODEL 4: FOURIER REGRESSION (captures multiple seasonal harmonics)
# ─────────────────────────────────────────────────────────────────────────────

def model_fourier_regression(y: np.ndarray, period: int, fp: int):
    n = len(y); period = max(period, 2)
    K = min(4, period // 2)  # number of Fourier pairs

    def build_X(idx_arr):
        t = np.array(idx_arr, dtype=float)
        cols = [np.ones(len(t)), t, t**2]  # intercept + trend + quadratic trend
        for k in range(1, K + 1):
            cols.append(np.sin(2 * np.pi * k * t / period))
            cols.append(np.cos(2 * np.pi * k * t / period))
        return np.column_stack(cols)

    X = build_X(range(n))
    # Ridge regression (L2 regularisation)
    lam = 1e4
    coef = np.linalg.solve(X.T @ X + lam * np.eye(X.shape[1]), X.T @ y)

    fitted = np.maximum(X @ coef, 0)
    X_fut  = build_X(range(n, n + fp))
    forecast = np.maximum(X_fut @ coef, 0).tolist()

    mape = float(np.mean(np.abs((y - fitted) / (np.abs(y) + 1e-9))) * 100)
    return {"name": "Fourier Harmonic Regression", "fitted": fitted.tolist(),
            "forecast": forecast, "mape": mape, "params": {"coef": coef.tolist()}}

# ─────────────────────────────────────────────────────────────────────────────
# MODEL 5: THETA METHOD (academic benchmark, often beats ARIMA)
# ─────────────────────────────────────────────────────────────────────────────

def model_theta(y: np.ndarray, period: int, fp: int):
    """
    Classical Theta method: decompose into 0-theta (straight line) and
    2-theta (amplified trend) lines, then average.
    """
    n = len(y)

    # Deseasonalise first
    _, seasonal, _ = stl_decompose(y, period=period, robust=False)
    y_sa = y - seasonal

    # Fit straight line (theta=0)
    t = np.arange(n, dtype=float)
    slope, intercept, *_ = stats.linregress(t, y_sa)
    theta0_fit = intercept + slope * t
    theta0_fc  = intercept + slope * (n + np.arange(1, fp + 1))

    # Fit theta=2 line (amplify curve from theta=0)
    # theta2 = 2*y - theta0
    theta2_series = 2 * y_sa - theta0_fit

    # Fit SES on theta2
    def ses_sse(alpha):
        alpha = float(alpha)
        if not (0 < alpha < 1): return 1e15
        L = theta2_series[0]; sse = 0.0
        for t_ in range(1, n):
            sse += (theta2_series[t_] - L)**2
            L = alpha * theta2_series[t_] + (1 - alpha) * L
        return sse

    from scipy.optimize import minimize_scalar
    r = minimize_scalar(ses_sse, bounds=(0.01, 0.99), method="bounded")
    alpha_opt = float(r.x)

    L = theta2_series[0]
    for t_ in range(1, n):
        L = alpha_opt * theta2_series[t_] + (1 - alpha_opt) * L
    theta2_fc = np.array([L] * fp)  # SES forecast is flat (last level)

    # Average theta0 and theta2 forecasts
    theta_fc_sa = (theta0_fc + theta2_fc) / 2

    # Add back seasonal
    seasonal_fc = np.array([seasonal[n - period + (h - 1) % period] for h in range(1, fp + 1)])
    forecast = np.maximum(theta_fc_sa + seasonal_fc, 0).tolist()

    # Fitted values
    theta0_vals = theta0_fit
    theta2_fitted = np.array([theta2_series[0]] + [
        alpha_opt * theta2_series[t_] + (1 - alpha_opt) *
        (alpha_opt * theta2_series[t_-1]) for t_ in range(1, n)
    ])
    fitted_sa = (theta0_vals + theta2_fitted[:n]) / 2
    fitted = np.maximum(fitted_sa + seasonal, 0)

    mape = float(np.mean(np.abs((y - fitted) / (np.abs(y) + 1e-9))) * 100)
    return {"name": "Theta Method", "fitted": fitted.tolist(),
            "forecast": forecast, "mape": mape, "params": {"alpha": alpha_opt}}

# ─────────────────────────────────────────────────────────────────────────────
# WALK-FORWARD VALIDATION (cross-validation on time series)
# ─────────────────────────────────────────────────────────────────────────────

def walk_forward_cv(y: np.ndarray, period: int, n_splits: int = 3):
    """
    Time-series cross-validation: train on prefix, test on next 1–3 months.
    Returns dict of model_name → average CV MAPE.
    """
    n = len(y)
    min_train = max(period * 2, 8)
    if n < min_train + 2:
        return {}

    step = max(1, (n - min_train) // (n_splits + 1))
    cv_errors = {}

    for split in range(n_splits):
        train_end = min_train + split * step
        if train_end >= n - 1: break
        test_end = min(train_end + max(1, period // 4), n)

        y_train = y[:train_end]
        y_test  = y[train_end:test_end]
        fp_cv   = len(y_test)

        for fn, nm in [
            (lambda y_, p, f: model_stl_damped(y_, p, f), "STL + Damped Holt"),
            (lambda y_, p, f: model_holtwinters(y_, p, f), "Holt-Winters ETS (opt.)"),
            (lambda y_, p, f: model_snaive_drift(y_, p, f), "Seasonal Naïve + Drift"),
            (lambda y_, p, f: model_fourier_regression(y_, p, f), "Fourier Harmonic Regression"),
            (lambda y_, p, f: model_theta(y_, p, f), "Theta Method"),
        ]:
            try:
                res = fn(y_train, period, fp_cv)
                fc  = np.array(res["forecast"][:fp_cv])
                err = float(np.mean(np.abs((y_test - fc) / (np.abs(y_test) + 1e-9))) * 100)
                cv_errors.setdefault(nm, []).append(err)
            except Exception:
                pass

    return {k: float(np.mean(v)) for k, v in cv_errors.items()}

# ─────────────────────────────────────────────────────────────────────────────
# RECENCY-WEIGHTED ENSEMBLE
# ─────────────────────────────────────────────────────────────────────────────

def ensemble_smart(models: list, y: np.ndarray, cv_mapes: dict, fp: int):
    """
    Weights based on:
      1. Walk-forward CV MAPE (if available)   → 60%
      2. In-sample MAPE on most recent 25%     → 40%
    Lower MAPE → higher weight.
    """
    n = len(y)
    recent_n = max(4, n // 4)
    recent_y = y[-recent_n:]

    scores = []
    for m in models:
        fitted_recent = np.array(m["fitted"][-recent_n:])
        recent_mape = float(np.mean(np.abs((recent_y - fitted_recent) / (np.abs(recent_y) + 1e-9))) * 100)
        cv_mape = cv_mapes.get(m["name"], m["mape"])  # fallback to in-sample if no CV
        combined = 0.6 * cv_mape + 0.4 * recent_mape
        scores.append(max(combined, 0.01))

    inv = np.array([1.0 / s for s in scores])
    weights = inv / inv.sum()

    blend_fc = sum(w * np.array(m["forecast"]) for w, m in zip(weights, models))
    blend_ft = sum(w * np.array(m["fitted"])   for w, m in zip(weights, models))

    # Ensemble MAPE (recent-weighted)
    ens_recent = sum(w * np.array(m["fitted"][-recent_n:]) for w, m in zip(weights, models))
    ens_mape = float(np.mean(np.abs((recent_y - ens_recent) / (np.abs(recent_y) + 1e-9))) * 100)

    return {
        "name": "⭐ Smart Ensemble",
        "fitted": blend_ft.tolist(),
        "forecast": blend_fc.tolist(),
        "mape": ens_mape,
        "params": {"weights": weights.tolist(), "scores": scores},
        "model_names": [m["name"] for m in models],
    }

# ─────────────────────────────────────────────────────────────────────────────
# CONFIDENCE INTERVALS  (empirical, not parametric)
# ─────────────────────────────────────────────────────────────────────────────

def conf_intervals(forecast: np.ndarray, residuals: np.ndarray, alpha: float = 0.10):
    """
    Uses empirical residual quantiles + horizon scaling.
    More honest than parametric normal assumption.
    """
    lo_q = np.percentile(residuals, alpha / 2 * 100)
    hi_q = np.percentile(residuals, (1 - alpha / 2) * 100)
    h = np.arange(1, len(forecast) + 1)
    scale = np.sqrt(h)
    lo = np.maximum(forecast + lo_q * scale, 0)
    hi = forecast + hi_q * scale
    return lo.tolist(), hi.tolist()

# ─────────────────────────────────────────────────────────────────────────────
# METRICS
# ─────────────────────────────────────────────────────────────────────────────

def compute_metrics(y_true, y_pred, name):
    y_true = np.array(y_true, dtype=float)
    y_pred = np.array(y_pred[:len(y_true)], dtype=float)
    mae  = np.mean(np.abs(y_true - y_pred))
    rmse = np.sqrt(np.mean((y_true - y_pred)**2))
    mape = np.mean(np.abs((y_true - y_pred) / (np.abs(y_true) + 1e-9))) * 100
    r2   = 1 - np.sum((y_true - y_pred)**2) / (np.sum((y_true - np.mean(y_true))**2) + 1e-9)
    return {"Model": name, "MAE (₹ Cr)": round(mae/1e7, 2),
            "RMSE (₹ Cr)": round(rmse/1e7, 2),
            "MAPE (%)": round(mape, 2), "R²": round(r2, 4)}

# ─────────────────────────────────────────────────────────────────────────────
# RUN ALL MODELS
# ─────────────────────────────────────────────────────────────────────────────

def run_all_models(y: np.ndarray, period: int, fp: int, progress_cb=None):
    models = []
    fns = [
        ("STL + Damped Holt",           model_stl_damped),
        ("Holt-Winters ETS (opt.)",      model_holtwinters),
        ("Seasonal Naïve + Drift",       model_snaive_drift),
        ("Fourier Harmonic Regression",  model_fourier_regression),
        ("Theta Method",                 model_theta),
    ]
    for i, (nm, fn) in enumerate(fns):
        try:
            m = fn(y, period, fp)
            models.append(m)
        except Exception as e:
            pass
        if progress_cb:
            progress_cb(int((i + 1) / (len(fns) + 2) * 100))

    if progress_cb: progress_cb(80)
    cv_mapes = {}
    try:
        cv_mapes = walk_forward_cv(y, period, n_splits=min(3, max(1, len(y) // 6)))
    except Exception:
        pass

    if progress_cb: progress_cb(90)
    ens = ensemble_smart(models, y, cv_mapes, fp)
    if progress_cb: progress_cb(100)
    return models, ens, cv_mapes

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(df, ens, models, future_dates, metrics_df, lo, hi, label, cv_mapes):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    thin = Border(*[Side(style="thin")] * 4)
    # thin border helper
    def apply_thin(cell):
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

    def hdr(cell, bg="1A2C5B"):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        apply_thin(cell)

    # Sheet 1: Forecast
    ws1 = wb.active; ws1.title = "6-Month Forecast"
    h1 = ["Month", "Ensemble Forecast (₹)", "Forecast (₹ Cr)", "Lower 90% CI", "Upper 90% CI", "MoM Change %"]
    for ci, h in enumerate(h1, 1):
        hdr(ws1.cell(1, ci, h))
        ws1.column_dimensions[get_column_letter(ci)].width = 24
    fc = np.array(ens["forecast"]); last_actual = df["Sales"].values[-1]
    for i, (fd, fv, fl, fh_v) in enumerate(zip(future_dates, fc, lo, hi)):
        prev = fc[i-1] if i > 0 else last_actual
        mom = (fv - prev) / (prev + 1e-9) * 100
        row = [pd.Timestamp(fd).strftime("%b %Y"), fv, fv/1e7, fl, fh_v, mom/100]
        fmts = ["@", "#,##0", "#,##0.00", "#,##0", "#,##0", '+0.0%;-0.0%']
        for ci, (v, fmt) in enumerate(zip(row, fmts), 1):
            c = ws1.cell(i+2, ci, v); c.number_format = fmt
            c.alignment = Alignment(horizontal="right" if ci > 1 else "center")
            apply_thin(c)

    # Sheet 2: Historical
    ws2 = wb.create_sheet("Historical + Fitted")
    h2 = ["Month", "Actual (₹)", "Actual (₹ Cr)", "Ensemble Fitted", "Residual", "MAPE Row %"]
    for ci, h in enumerate(h2, 1):
        hdr(ws2.cell(1, ci, h), "0D7377")
        ws2.column_dimensions[get_column_letter(ci)].width = 22
    fitted = np.array(ens["fitted"])
    for i, (dt, act) in enumerate(zip(df["Date"], df["Sales"])):
        res_ = act - fitted[i]; mr = abs(res_)/(act+1e-9)*100
        row = [dt.strftime("%b %Y"), act, act/1e7, fitted[i], res_, mr]
        fmts = ["@", "#,##0", "#,##0.00", "#,##0", "#,##0", "0.00"]
        for ci, (v, fmt) in enumerate(zip(row, fmts), 1):
            c = ws2.cell(i+2, ci, v); c.number_format = fmt
            c.alignment = Alignment(horizontal="right" if ci > 1 else "center")
            apply_thin(c)

    # Sheet 3: Model Metrics
    ws3 = wb.create_sheet("Model Metrics")
    for ci, h in enumerate(metrics_df.columns, 1):
        hdr(ws3.cell(1, ci, h), "4A148C")
        ws3.column_dimensions[get_column_letter(ci)].width = 26
    for ri, row_data in enumerate(metrics_df.itertuples(index=False), 2):
        for ci, val in enumerate(row_data, 1):
            c = ws3.cell(ri, ci, val)
            if isinstance(val, float): c.number_format = "0.00"
            c.alignment = Alignment(horizontal="center")
            apply_thin(c)

    # Sheet 4: CV Results
    ws4 = wb.create_sheet("Cross-Validation")
    hdr(ws4.cell(1, 1, "Model"), "6A1B9A"); hdr(ws4.cell(1, 2, "CV MAPE (%)"), "6A1B9A")
    ws4.column_dimensions["A"].width = 32; ws4.column_dimensions["B"].width = 20
    for i, (nm, cv_m) in enumerate(cv_mapes.items(), 2):
        c1 = ws4.cell(i, 1, nm); c2 = ws4.cell(i, 2, round(cv_m, 2))
        c2.number_format = "0.00"
        for c in [c1, c2]: apply_thin(c); c.alignment = Alignment(horizontal="center")

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# CHARTS
# ─────────────────────────────────────────────────────────────────────────────

def fig_main(df, models, ens, future_dates, lo, hi):
    fig = go.Figure()
    y_cr = df["Sales"] / 1e7
    fig.add_trace(go.Scatter(x=df["Date"], y=y_cr, mode="lines+markers",
                             name="Actual", line=dict(color="#ffffff", width=2.5),
                             marker=dict(size=5)))
    for m, c in zip(models, ACCENT):
        fig.add_trace(go.Scatter(x=future_dates, y=np.array(m["forecast"])/1e7,
                                 mode="lines", name=f"{m['name']} ({m['mape']:.1f}%)",
                                 line=dict(color=c, width=1.5, dash="dot"), opacity=0.7))
    fig.add_trace(go.Scatter(
        x=list(future_dates) + list(future_dates)[::-1],
        y=list(np.array(hi)/1e7) + list(np.array(lo)/1e7)[::-1],
        fill="toself", fillcolor="rgba(230,80,150,0.12)",
        line=dict(color="rgba(0,0,0,0)"), name="90% CI", hoverinfo="skip"))
    fig.add_trace(go.Scatter(x=future_dates, y=np.array(ens["forecast"])/1e7,
                             mode="lines+markers",
                             name=f"⭐ Smart Ensemble ({ens['mape']:.1f}% recent MAPE)",
                             line=dict(color="#e650a0", width=3),
                             marker=dict(size=9, symbol="star")))
    fig.add_vline(x=str(df["Date"].max()), line_dash="dash",
                  line_color="rgba(255,255,255,0.3)", line_width=1.5)
    fig.update_layout(title="📈 All Models + Smart Ensemble Forecast",
                      xaxis_title="Month", yaxis_title="Sales (₹ Crore)",
                      legend=dict(orientation="h", yanchor="bottom", y=-0.32, x=0),
                      height=490, **PLOTLY_THEME)
    return fig


def fig_components(stl_model, df):
    t = stl_model["components"]["trend"]
    s = stl_model["components"]["seasonal"]
    r = stl_model["components"]["residual"]
    dates = df["Date"].values

    fig = make_subplots(rows=3, cols=1, shared_xaxes=True,
                        subplot_titles=("Trend (₹ Cr)", "Seasonal Component", "Residual"),
                        vertical_spacing=0.1)
    fig.add_trace(go.Scatter(x=dates, y=t/1e7, line=dict(color="#7c6fee", width=2),
                             name="Trend"), row=1, col=1)
    fig.add_trace(go.Scatter(x=dates, y=s/1e7, line=dict(color="#f5a623", width=1.5),
                             name="Seasonal"), row=2, col=1)
    fig.add_trace(go.Bar(x=dates, y=r/1e7, name="Residual",
                         marker_color=np.where(r >= 0, "#50d090", "#f55050")),
                  row=3, col=1)
    fig.update_layout(height=520, title="🔬 STL Decomposition — Trend / Seasonal / Residual",
                      showlegend=False, **PLOTLY_THEME)
    return fig


def fig_cv(cv_mapes):
    names = list(cv_mapes.keys())
    vals  = list(cv_mapes.values())
    colors = [ACCENT[i % len(ACCENT)] for i in range(len(names))]
    fig = go.Figure(go.Bar(x=vals, y=names, orientation="h",
                           marker=dict(color=colors, opacity=0.85),
                           text=[f"{v:.1f}%" for v in vals], textposition="outside"))
    fig.add_vline(x=5,  line_dash="dash", line_color="rgba(80,200,120,0.5)")
    fig.add_vline(x=10, line_dash="dash", line_color="rgba(245,166,35,0.5)")
    fig.update_layout(title="🎯 Walk-Forward Cross-Validation MAPE",
                      xaxis_title="CV MAPE %", height=340, **PLOTLY_THEME)
    return fig


def fig_outliers(raw, clean, flags, dates):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=dates, y=raw/1e7, mode="lines",
                             name="Raw", line=dict(color="rgba(150,150,200,0.5)", dash="dot")))
    fig.add_trace(go.Scatter(x=dates, y=clean/1e7, mode="lines+markers",
                             name="Cleaned", line=dict(color="#50d090", width=2.5),
                             marker=dict(size=5)))
    ol_d = [d for d, f in zip(dates, flags) if f]
    ol_v = [v for v, f in zip(raw/1e7, flags) if f]
    if ol_d:
        fig.add_trace(go.Scatter(x=ol_d, y=ol_v, mode="markers", name="Outlier",
                                 marker=dict(size=12, color="#ff5050", symbol="x",
                                             line=dict(width=2))))
    fig.update_layout(title="🔍 Outlier Detection (Hampel Identifier)",
                      xaxis_title="Month", yaxis_title="₹ Crore", height=360, **PLOTLY_THEME)
    return fig


def fig_seasonality(df):
    d = df.copy(); d["Month"] = d["Date"].dt.month; d["Year"] = d["Date"].dt.year
    pivot = d.pivot_table(values="Sales", index="Year", columns="Month", aggfunc="sum") / 1e7
    mn = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    cl = [mn[c-1] for c in pivot.columns]
    fig = go.Figure(go.Heatmap(z=pivot.values, x=cl, y=[str(y) for y in pivot.index],
                               colorscale="YlOrRd",
                               text=np.round(pivot.values, 1),
                               texttemplate="₹%{text}Cr",
                               textfont=dict(size=9),
                               hovertemplate="Year:%{y} Month:%{x} Sales:₹%{z:.1f}Cr<extra></extra>"))
    fig.update_layout(title="🗓️ Seasonality Heatmap (₹ Crore)",
                      height=320, **PLOTLY_THEME)
    return fig


def fig_detailed(df, ens, future_dates, lo, hi):
    tail = df.tail(6)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=tail["Date"], y=tail["Sales"]/1e7,
                             mode="lines+markers", name="Last 6M Actual",
                             line=dict(color="#aaaaee", width=2)))
    fc_cr = np.array(ens["forecast"])/1e7
    lo_cr = np.array(lo)/1e7; hi_cr = np.array(hi)/1e7
    fig.add_trace(go.Scatter(
        x=list(future_dates)+list(future_dates)[::-1],
        y=list(hi_cr)+list(lo_cr)[::-1],
        fill="toself", fillcolor="rgba(230,80,150,0.15)",
        line=dict(color="rgba(0,0,0,0)"), name="90% CI", hoverinfo="skip"))
    fig.add_trace(go.Scatter(x=future_dates, y=fc_cr,
                             mode="lines+markers+text",
                             name="⭐ Ensemble Forecast",
                             line=dict(color="#e650a0", width=3),
                             marker=dict(size=10, symbol="star"),
                             text=[f"₹{v:.1f}Cr" for v in fc_cr],
                             textposition="top center",
                             textfont=dict(size=11, color="#f080c0")))
    fig.update_layout(title="📅 Detailed 6-Month Forecast with 90% CI",
                      xaxis_title="Month", yaxis_title="₹ Crore",
                      height=420, **PLOTLY_THEME)
    return fig

# ─────────────────────────────────────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.markdown("""
    <div style="text-align:center; padding:20px 0 10px 0;">
        <div style="font-size:40px; font-weight:800;
             background:linear-gradient(135deg,#a09cf7,#f5a623,#50d090);
             -webkit-background-clip:text; -webkit-text-fill-color:transparent;">
            📈 Textile Sales Forecasting Pro
        </div>
        <div style="color:#7777aa; font-size:14px; margin-top:8px; letter-spacing:1px;">
            STL Decomposition · Theta Method · Fourier Regression · Walk-Forward CV · Smart Ensemble
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    # ── Sidebar ───────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### ⚙️ Settings")
        st.markdown("---")
        uploaded = st.file_uploader("📂 Upload Excel File", type=["xlsx","xls"],
                                    help="Columns: Date · CY Value · SUBGRP")
        subgrp_choice = None; remove_outliers = True; fp = 6

        if uploaded:
            st.success("✅ File uploaded!")
            try:
                raw_df = load_excel(uploaded.read())
                col_map = find_columns(raw_df)
                df_r = dedup_cols(raw_df.rename(columns=col_map).copy())
                if "SUBGRP" in df_r.columns:
                    sg = safe_series(df_r, "SUBGRP")
                    sgs = sorted(sg.dropna().unique().tolist())
                    subgrp_choice = st.selectbox("🏷️ Filter by SUBGRP",
                                                 ["All Products"] + sgs, index=0)
                else:
                    st.warning("No SUBGRP column — using all data.")
                st.markdown("---")
                remove_outliers = st.toggle("🔍 Remove Outliers (Hampel)", value=True)
                fp = st.slider("📅 Forecast Months", 3, 12, 6)
                st.markdown("---")
                st.markdown("**File Info**")
                st.caption(f"Rows: {len(raw_df):,}")
                st.caption(f"Cols: {list(raw_df.columns)}")
            except Exception as e:
                st.error(f"❌ Error: {e}"); return

        st.markdown("---")
        st.markdown("""<div style="font-size:11px;color:#555588;text-align:center;">
        Expected: <code>Date · CY Value · SUBGRP</code><br><br>
        <b>💡 Tip:</b> Upload ≥18 months for full accuracy
        </div>""", unsafe_allow_html=True)

    if uploaded is None:
        c1, c2, c3 = st.columns(3)
        for col, icon, title, desc in [
            (c1, "📂", "Upload Excel", "Date · CY Value · SUBGRP columns"),
            (c2, "🏷️", "Pick SUBGRP", "Forecast any product individually"),
            (c3, "🔮", "5-Model Forecast", "STL · HW · Theta · Fourier · Naïve + Smart Ensemble"),
        ]:
            with col:
                st.markdown(f"""<div class="metric-card" style="text-align:center;padding:32px 20px;">
                    <div style="font-size:36px;">{icon}</div>
                    <div style="font-size:17px;font-weight:700;color:#a09cf7;margin:10px 0 8px;">{title}</div>
                    <div style="font-size:13px;color:#666699;">{desc}</div></div>""",
                    unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("""<div class="banner-warn">
            <div class="banner-title warn-title">📏 Data Length Guide</div>
            <div class="banner-body warn-body">
            &lt;6 months → ❌ Cannot forecast<br>
            6–17 months → ⚠️ Limited accuracy<br>
            ≥18 months → ✅ Full accuracy (recommended)
            </div></div>""", unsafe_allow_html=True)
        st.info("👈 Upload your Excel file from the sidebar.")
        return

    # ── Process data ──────────────────────────────────────────────────────
    try:
        raw_df  = load_excel(uploaded.getvalue())
        col_map = find_columns(raw_df)
        if "Date"  not in col_map.values(): st.error("❌ 'Date' column missing!"); return
        if "Sales" not in col_map.values(): st.error("❌ 'CY Value'/'Sales' column missing!"); return

        df_work = dedup_cols(raw_df.rename(columns=col_map).copy())
        df_work["Date"]  = pd.to_datetime(safe_series(df_work,"Date"), errors="coerce")
        df_work["Sales"] = pd.to_numeric(safe_series(df_work,"Sales"), errors="coerce").fillna(0)
        df_work = df_work.dropna(subset=["Date","Sales"])

        label = "All Products"
        if subgrp_choice and subgrp_choice != "All Products" and "SUBGRP" in df_work.columns:
            df_work = df_work[safe_series(df_work,"SUBGRP") == subgrp_choice].copy()
            label = subgrp_choice

        if len(df_work) == 0: st.error("❌ No data after filtering!"); return

        df_m = (df_work.groupby(df_work["Date"].dt.to_period("M"))["Sales"]
                .sum().reset_index())
        df_m["Date"] = df_m["Date"].dt.to_timestamp()
        df_m = df_m.sort_values("Date").reset_index(drop=True)

        if len(df_m) < MIN_MONTHS_BASIC:
            st.error(f"❌ Only {len(df_m)} months — need ≥{MIN_MONTHS_BASIC}."); return

        ds = data_status(len(df_m))
        raw_s = df_m["Sales"].copy()
        cleaned_s, flags, ol_info = detect_outliers(df_m["Sales"])
        df_fc = df_m.copy()
        if remove_outliers:
            df_fc["Sales"] = cleaned_s

    except Exception as e:
        st.error(f"❌ Error processing: {e}")
        import traceback; st.code(traceback.format_exc()); return

    # ── Data-length banner ────────────────────────────────────────────────
    cls = "banner-good" if ds["tier"] == "full" else "banner-warn"
    tc  = "good-title"  if ds["tier"] == "full" else "warn-title"
    bc  = "good-body"   if ds["tier"] == "full" else "warn-body"
    st.markdown(f"""<div class="{cls}">
        <div class="banner-title {tc}">📊 Data Status — {len(df_m)} months loaded</div>
        <div class="banner-body {bc}">{ds['msg']}</div></div>""", unsafe_allow_html=True)

    # ── KPI Row ───────────────────────────────────────────────────────────
    st.markdown(f"<div class='section-header'>📊 Overview — {label}</div>", unsafe_allow_html=True)
    k1,k2,k3,k4,k5 = st.columns(5)
    avg_cr  = df_fc["Sales"].mean()/1e7
    peak_cr = df_fc["Sales"].max()/1e7
    peak_m  = df_fc.loc[df_fc["Sales"].idxmax(),"Date"].strftime("%b %Y")
    total_cr = df_fc["Sales"].sum()/1e7
    df_fc["Year"] = df_fc["Date"].dt.year
    y_last = df_fc["Year"].max(); y_prev = y_last - 1
    s_last = df_fc[df_fc["Year"]==y_last]["Sales"].sum()
    s_prev = df_fc[df_fc["Year"]==y_prev]["Sales"].sum()
    yoy = (s_last - s_prev)/(s_prev+1e-9)*100 if s_prev > 0 else 0
    period_str = f"{df_m['Date'].min().strftime('%b %Y')} → {df_m['Date'].max().strftime('%b %Y')}"
    for col, lbl, val, sub in [
        (k1, "PERIOD",      period_str,         f"{len(df_m)} months"),
        (k2, "AVG MONTHLY", f"₹{avg_cr:.1f} Cr","avg"),
        (k3, "PEAK MONTH",  f"₹{peak_cr:.1f} Cr", peak_m),
        (k4, "TOTAL",       f"₹{total_cr:.0f} Cr","all months"),
        (k5, "YoY GROWTH",  f"{'🔺' if yoy>0 else '🔻'}{abs(yoy):.1f}%", f"{y_prev}→{y_last}"),
    ]:
        with col:
            st.markdown(f"""<div class="metric-card">
                <div class="metric-label">{lbl}</div>
                <div class="metric-value" style="font-size:{'17px' if len(val)>9 else '22px'};">{val}</div>
                <div class="metric-sub">{sub}</div></div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    if ol_info["n_outliers"] > 0:
        badge = "✅ Outliers cleaned" if remove_outliers else "⚠️ Outliers detected (not cleaned)"
        st.markdown(f'<span class="{"acc-high" if remove_outliers else "acc-mid"}">{badge} — {ol_info["n_outliers"]} points</span>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<span class="acc-high">✅ No outliers — data is clean</span>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    run_col, _ = st.columns([1, 3])
    with run_col:
        run_btn = st.button("🚀 Run Forecast", use_container_width=True)

    if "fc_done" not in st.session_state: st.session_state.fc_done = False
    if run_btn:
        st.session_state.fc_done = True
        st.session_state.fc_key  = f"{label}_{remove_outliers}_{fp}"

    if not st.session_state.fc_done:
        st.info("👆 Click **Run Forecast** to generate predictions.")
        st.plotly_chart(fig_outliers(raw_s, cleaned_s, flags, df_m["Date"]), use_container_width=True)
        return

    # ── Train ─────────────────────────────────────────────────────────────
    y = df_fc["Sales"].values
    with st.spinner("🔧 Training 5 models + walk-forward CV + smart ensemble..."):
        prog = st.progress(0)
        models, ens, cv_mapes = run_all_models(y, ds["sl"], fp, progress_cb=prog.progress)

    last_date  = df_fc["Date"].max()
    future_dates = pd.date_range(start=last_date + pd.offsets.MonthBegin(1), periods=fp, freq="MS")
    residuals  = y - np.array(ens["fitted"])
    lo, hi_ci  = conf_intervals(np.array(ens["forecast"]), residuals)
    all_models = models + [ens]
    metrics_df = pd.DataFrame([compute_metrics(y, m["fitted"], m["name"]) for m in all_models])

    # Find STL model for component plot
    stl_model = next((m for m in models if "STL" in m["name"]), None)

    # ── Tabs ──────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🔮 Forecast", "🔬 Decomposition", "📊 Data Quality", "🎯 CV Accuracy", "📋 Metrics"])

    with tab1:
        st.markdown("<div class='section-header'>🔮 Forecast Results</div>", unsafe_allow_html=True)
        st.markdown(acc_badge(ens["mape"]), unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)

        fc_arr = np.array(ens["forecast"]); total_fc = fc_arr.sum()/1e7
        best_idx = np.argmax(fc_arr); best_val = fc_arr[best_idx]/1e7
        last_act = y[-1]; growth = (fc_arr[0]-last_act)/(last_act+1e-9)*100

        ka,kb,kc,kd = st.columns(4)
        for col, lbl, val, sub in [
            (ka, "6M TOTAL FORECAST",    f"₹{total_fc:.1f} Cr",   "ensemble"),
            (kb, "RECENT MAPE",          f"{ens['mape']:.1f}%",    "last 25% of data"),
            (kc, "PEAK FORECAST",        f"₹{best_val:.1f} Cr",   future_dates[best_idx].strftime("%b %Y")),
            (kd, "NEXT MONTH CHANGE",    f"{growth:+.1f}%",        "vs last actual"),
        ]:
            with col:
                clr = "#50d090" if lbl == "RECENT MAPE" and ens["mape"] < 10 else "#a09cf7"
                st.markdown(f"""<div class="metric-card">
                    <div class="metric-label">{lbl}</div>
                    <div class="metric-value" style="color:{clr};">{val}</div>
                    <div class="metric-sub">{sub}</div></div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.plotly_chart(fig_main(df_fc, models, ens, future_dates, lo, hi_ci), use_container_width=True)
        st.plotly_chart(fig_detailed(df_fc, ens, future_dates, lo, hi_ci), use_container_width=True)

        st.markdown("<div class='section-header'>📋 Forecast Table</div>", unsafe_allow_html=True)
        last_actual = y[-1]
        fc_rows = []
        for i, (fd, fv, fl, fh_v) in enumerate(zip(future_dates, fc_arr, lo, hi_ci)):
            prev = fc_arr[i-1] if i > 0 else last_actual
            mom = (fv - prev)/(prev+1e-9)*100
            fc_rows.append({
                "Month":            fd.strftime("%b %Y"),
                "Forecast (₹ Cr)":  f"₹{fv/1e7:.2f} Cr",
                "Lower CI":         f"₹{fl/1e7:.1f} Cr",
                "Upper CI":         f"₹{fh_v/1e7:.1f} Cr",
                "MoM Change":       f"{'🔺' if mom>0 else '🔻'} {mom:+.1f}%",
            })
        st.dataframe(pd.DataFrame(fc_rows), use_container_width=True, hide_index=True)

        excel_b = build_excel(df_fc, ens, models, future_dates,
                              metrics_df, lo, hi_ci, label, cv_mapes)
        st.download_button("⬇️ Download Full Results (Excel)", data=excel_b,
                           file_name=f"forecast_{label.replace(' ','_')}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)

    with tab2:
        st.markdown("<div class='section-header'>🔬 STL Decomposition</div>", unsafe_allow_html=True)
        if stl_model and "components" in stl_model:
            st.plotly_chart(fig_components(stl_model, df_fc), use_container_width=True)
            comps = stl_model["components"]
            trend_slope = (comps["trend"][-1] - comps["trend"][0]) / max(len(comps["trend"])-1, 1)
            trend_dir   = "📈 Upward" if trend_slope > 0 else "📉 Downward"
            seas_range  = comps["seasonal"].max() - comps["seasonal"].min()
            st.markdown(f"""<div class="banner-good">
                <div class="banner-title good-title">🔍 Decomposition Insights</div>
                <div class="banner-body good-body">
                <b>Trend direction:</b> {trend_dir} (slope ₹{trend_slope/1e7:.2f} Cr/month)<br>
                <b>Seasonal swing:</b> ₹{seas_range/1e7:.2f} Cr peak-to-trough<br>
                <b>Residual std:</b> ₹{np.std(comps['residual'])/1e7:.2f} Cr (lower = cleaner data)
                </div></div>""", unsafe_allow_html=True)
        else:
            st.info("Decomposition not available for this data length.")

    with tab3:
        st.markdown("<div class='section-header'>📊 Data Quality</div>", unsafe_allow_html=True)
        st.plotly_chart(fig_outliers(raw_s, cleaned_s, flags, df_m["Date"]), use_container_width=True)
        st.plotly_chart(fig_seasonality(df_fc), use_container_width=True)

    with tab4:
        st.markdown("<div class='section-header'>🎯 Walk-Forward Cross-Validation</div>", unsafe_allow_html=True)
        if cv_mapes:
            st.plotly_chart(fig_cv(cv_mapes), use_container_width=True)
            cv_df = pd.DataFrame([{"Model": k, "CV MAPE (%)": round(v, 2)}
                                   for k, v in cv_mapes.items()])
            st.dataframe(cv_df, use_container_width=True, hide_index=True)
            st.markdown("""
            **Walk-forward CV**: trains on historical data, tests on unseen future months —
            gives the most honest estimate of real-world accuracy (not overfitted in-sample fit).
            """)
        else:
            st.info("Need more data for cross-validation. Upload ≥18 months for best results.")

        # Model weight breakdown
        st.markdown("**Smart Ensemble Weights:**")
        wts = ens["params"].get("weights", [])
        w_rows = [{"Model": nm, "Weight %": f"{w*100:.1f}%",
                   "Stars": "⭐⭐⭐" if w == max(wts) else ("⭐⭐" if w > 0.25 else "⭐")}
                  for nm, w in zip(ens.get("model_names", [m["name"] for m in models]), wts)]
        st.dataframe(pd.DataFrame(w_rows), use_container_width=True, hide_index=True)

    with tab5:
        st.markdown("<div class='section-header'>📋 Full Model Metrics</div>", unsafe_allow_html=True)
        if ds["tier"] != "full":
            st.markdown(f"""<div class="banner-warn">
                <div class="banner-title warn-title">⚠️ Limited data — accuracy may be lower</div>
                <div class="banner-body warn-body">Upload {MIN_MONTHS_FULL - len(df_m)} more months for full accuracy.</div>
            </div>""", unsafe_allow_html=True)
        st.dataframe(
            metrics_df.style
                .format({"MAE (₹ Cr)":"{:.2f}","RMSE (₹ Cr)":"{:.2f}","MAPE (%)":"{:.2f}","R²":"{:.4f}"})
                .background_gradient(subset=["MAPE (%)"], cmap="RdYlGn_r"),
            use_container_width=True, hide_index=True)
        st.markdown("""
        **Metrics Guide:**
        - **MAPE** — Mean Absolute % Error (lower = better). Ensemble uses *recent-period MAPE*, not full-history.
        - **R²** — Fit quality; 1.0 = perfect.
        - **MAE / RMSE** — Error in ₹ Crore. RMSE penalises large errors more.
        - **Ensemble weight** is based on CV MAPE (60%) + recent in-sample MAPE (40%).
        """)


if __name__ == "__main__":
    main()
